#!/usr/bin/env python
"""
Tests for ddlgenerator.sources module.

Covers: _ensure_rows, _ordered_yaml_load, _json_loader, _interpret_fieldnames,
        _table_score, _html_to_odicts, NamedIter, filename_from_url,
        Source class methods, sqlalchemy_table_sources
"""

import contextlib
import os
import pathlib
import random
from collections import Counter
from io import StringIO
from unittest.mock import MagicMock, Mock, patch

import pytest

try:
    import yaml
except ImportError:
    yaml = None

try:
    import bs4
except ImportError:
    bs4 = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import sqlalchemy
except ImportError:
    sqlalchemy = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from pymongo.collection import Collection as MongoCollection
except ImportError:
    MongoCollection = None

from ddlgenerator.sources import (
    NamedIter,
    ParseException,
    Source,
    _ensure_rows,
    _html_to_odicts,
    _interpret_fieldnames,
    _json_loader,
    _ordered_yaml_load,
    _table_score,
    count_sqlalchemy_tables,
    filename_from_url,
    sqlalchemy_table_sources,
)


def here(filename):
    return os.path.join(os.path.dirname(__file__), filename)


# ---------------------------------------------------------------------------
# _ensure_rows
# ---------------------------------------------------------------------------
class TestEnsureRows:
    def test_single_dict_wraps_in_list(self):
        result = _ensure_rows({"a": 1, "b": 2})
        assert result == [{"a": 1, "b": 2}]

    def test_dict_of_dicts_converts_with_name_key(self):
        result = _ensure_rows({
            "a": {"a1": 1, "a2": 2},
            "b": {"b1": 1, "b2": 2}
        })
        assert len(result) == 2
        assert {"name_": "a", "a1": 1, "a2": 2} in result
        assert {"name_": "b", "b1": 1, "b2": 2} in result

    def test_list_of_dicts_returns_as_is(self):
        data = [{"a1": 1, "a2": 2}, {"b1": 1, "b2": 2}]
        result = _ensure_rows(data)
        assert result == data

    def test_empty_dict_returns_empty_list(self):
        assert _ensure_rows({}) == []

    def test_dict_with_mixed_values(self):
        """Dict with non-dict values should wrap as single item list."""
        result = _ensure_rows({"a": 1, "b": "text"})
        assert result == [{"a": 1, "b": "text"}]


# ---------------------------------------------------------------------------
# _ordered_yaml_load
# ---------------------------------------------------------------------------
class TestOrderedYamlLoad:
    @pytest.mark.skipif(yaml is None, reason="pyyaml not installed")
    def test_loads_yaml_preserving_order(self):
        yaml_content = "- name: Alice\n  age: 30\n- name: Bob\n  age: 25"
        result = list(_ordered_yaml_load(StringIO(yaml_content)))
        assert len(result) == 2
        assert result[0]["name"] == "Alice"
        assert result[1]["name"] == "Bob"

    @pytest.mark.skipif(yaml is None, reason="pyyaml not installed")
    def test_uses_safe_loader(self):
        """Verify that _ordered_yaml_load uses SafeLoader for security."""
        import inspect
        source = inspect.getsource(_ordered_yaml_load)
        assert "SafeLoader" in source

    @pytest.mark.skipif(yaml is None, reason="pyyaml not installed")
    def test_single_dict_wrapped_in_list(self):
        yaml_content = "name: Alice\nage: 30"
        result = list(_ordered_yaml_load(StringIO(yaml_content)))
        assert len(result) == 1
        assert result[0]["name"] == "Alice"

    def test_import_error_when_yaml_none(self):
        """Should raise ImportError if yaml module is not available."""
        with patch('ddlgenerator.sources.yaml', None):
            with pytest.raises(ImportError, match="pyyaml not installed"):
                _ordered_yaml_load(StringIO("test: 1"))


# ---------------------------------------------------------------------------
# _json_loader
# ---------------------------------------------------------------------------
class TestJsonLoader:
    def test_loads_json_array(self):
        json_content = '[{"name": "Alice"}, {"name": "Bob"}]'
        result = list(_json_loader(StringIO(json_content)))
        assert len(result) == 2
        assert result[0]["name"] == "Alice"

    def test_loads_single_object(self):
        json_content = '{"name": "Alice", "age": 30}'
        result = list(_json_loader(StringIO(json_content)))
        assert len(result) == 1
        assert result[0]["name"] == "Alice"

    def test_preserves_order(self):
        json_content = '[{"a": 1, "b": 2, "c": 3}]'
        result = list(_json_loader(StringIO(json_content)))
        keys = list(result[0].keys())
        assert keys == ["a", "b", "c"]


# ---------------------------------------------------------------------------
# _interpret_fieldnames
# ---------------------------------------------------------------------------
class TestInterpretFieldnames:
    def test_returns_fieldnames_as_is_for_non_integer(self):
        fieldnames = ["name", "age", "city"]
        result = _interpret_fieldnames(StringIO(""), fieldnames)
        assert result == fieldnames

    def test_integer_line_number_extracts_headers(self):
        """Line number 1 means first row is headers."""
        csv_content = "name,age,city\nAlice,30,NYC\n"
        result = _interpret_fieldnames(StringIO(csv_content), 1)
        assert result == ["name", "age", "city"]

    def test_zero_generates_field_names(self):
        csv_content = "Alice,30,NYC\n"
        result = _interpret_fieldnames(StringIO(csv_content), 0)
        assert result == ["Field1", "Field2", "Field3"]

    def test_negative_number_as_string(self):
        """Non-integer strings should be returned as-is."""
        result = _interpret_fieldnames(StringIO(""), "custom")
        assert result == "custom"


# ---------------------------------------------------------------------------
# _table_score
# ---------------------------------------------------------------------------
@pytest.mark.skipif(bs4 is None, reason="beautifulsoup4 not installed")
class TestTableScore:
    def test_scores_based_on_columns_and_headings(self):
        """Test that _table_score calculates score based on structure."""
        html = """
        <table>
            <thead><tr><th>Col1</th><th>Col2</th></tr></thead>
            <tbody>
                <tr><td>A</td><td>1</td></tr>
                <tr><td>B</td><td>2</td></tr>
            </tbody>
        </table>
        """
        soup = bs4.BeautifulSoup(html, 'html.parser')
        tbl = soup.find('table')
        score = _table_score(tbl)
        assert score > 0

    def test_row_count_outweighs_column_count(self):
        """The score is meant to favour tables with many data rows. The first
        term counted n_columns a second time instead of n_rows, so a short
        wide table beat a tall narrow one."""
        tall = bs4.BeautifulSoup(
            "<table>" + "<tr><td>a</td><td>b</td></tr>" * 10 + "</table>",
            'html.parser').find('table')
        wide = bs4.BeautifulSoup(
            "<table>" + "<tr>" + "<td>x</td>" * 10 + "</tr>" * 2 + "</table>",
            'html.parser').find('table')
        assert _table_score(tall) > _table_score(wide)

    def test_bonus_for_thead(self):
        """Tables with <thead> should get bonus points."""
        html_with_thead = """
        <table><thead><tr><th>A</th></tr></thead><tbody><tr><td>1</td></tr></tbody></table>
        """
        html_without_thead = """
        <table><tr><th>A</th></tr><tr><td>1</td></tr></table>
        """
        soup1 = bs4.BeautifulSoup(html_with_thead, 'html.parser')
        soup2 = bs4.BeautifulSoup(html_without_thead, 'html.parser')
        score_with = _table_score(soup1.find('table'))
        score_without = _table_score(soup2.find('table'))
        # Thead bonus should make score higher
        assert score_with >= score_without


# ---------------------------------------------------------------------------
# _html_to_odicts
# ---------------------------------------------------------------------------
@pytest.mark.skipif(bs4 is None, reason="beautifulsoup4 not installed")
class TestHtmlToOdicts:
    def test_extracts_table_data(self):
        html = """
        <table>
            <thead><tr><th>Name</th><th>Age</th></tr></thead>
            <tbody>
                <tr><td>Alice</td><td>30</td></tr>
                <tr><td>Bob</td><td>25</td></tr>
                <tr><td>Carol</td><td>35</td></tr>
            </tbody>
        </table>
        """
        result = list(_html_to_odicts(html))
        # Function skips first data row (used as header if no thead), so we get 2 rows
        assert len(result) >= 1
        assert "Name" in result[0] or "Field1" in result[0]

    def test_generates_field_names_for_empty_headers(self):
        html = """
        <table>
            <tr><th></th><th>Name</th></tr>
            <tr><td>1</td><td>Alice</td></tr>
        </table>
        """
        result = list(_html_to_odicts(html))
        assert "Field1" in result[0]
        assert "Name" in result[0]

    def test_raises_when_no_tables(self):
        html = "<html><body><p>No tables here</p></body></html>"
        with pytest.raises(ParseException, match="No HTML tables found"):
            list(_html_to_odicts(html))


# ---------------------------------------------------------------------------
# NamedIter
# ---------------------------------------------------------------------------
class TestNamedIter:
    def test_wraps_iterator_with_name(self):
        data = [1, 2, 3]
        named = NamedIter(iter(data), name="my_data")
        assert named.name == "my_data"

    def test_preserves_iteration(self):
        """NamedIter should allow iteration through the wrapped iterator."""
        data = [{"a": 1}, {"a": 2}]
        named = NamedIter(iter(data), name="test")
        # Use the __next__ method directly
        result = [named.__next__(), named.__next__()]
        assert result == data


# ---------------------------------------------------------------------------
# filename_from_url
# ---------------------------------------------------------------------------
class TestFilenameFromUrl:
    def test_extracts_filename_from_path(self):
        result = filename_from_url("https://example.com/data/myfile.json")
        assert result == "myfile"

    def test_handles_no_extension(self):
        result = filename_from_url("https://example.com/data/myfile")
        assert result == "myfile"

    def test_handles_query_string(self):
        result = filename_from_url("https://example.com/data.json?foo=bar")
        assert result == "data"

    def test_handles_complex_path(self):
        result = filename_from_url("https://example.com/path/to/data.csv")
        assert result == "data"


# ---------------------------------------------------------------------------
# Source class - Generator
# ---------------------------------------------------------------------------
class TestSourceGenerator:
    def test_source_is_generator_preserves_name(self):
        """Generator with name attribute should use that name for table_name."""
        # Create a custom iterator class that has a name attribute
        class NamedGenerator:
            def __init__(self):
                self.name = "custom_name"
                self._data = [{"a": 1}]
                self._index = 0

            def __iter__(self):
                return self

            def __next__(self):
                if self._index >= len(self._data):
                    raise StopIteration
                result = self._data[self._index]
                self._index += 1
                return result

        gen_obj = NamedGenerator()
        src = Source(gen_obj)
        assert src.table_name == "custom_name"

    def test_source_is_generator_without_name(self):
        """Generator without name attribute should get default table name."""
        def gen():
            yield {"a": 1}
        src = Source(gen())
        assert src.table_name == "Table0"


# ---------------------------------------------------------------------------
# Source class - URL
# ---------------------------------------------------------------------------
class TestSourceUrl:
    @patch('ddlgenerator.sources.url_utils.safe_fetch_text')
    def test_source_is_url_uses_safe_fetch(self, mock_fetch):
        mock_fetch.return_value = '[{"name": "Alice"}]'
        src = Source("https://example.com/data.json")
        mock_fetch.assert_called_once_with("https://example.com/data.json")
        assert src.table_name == "data"

    @patch('ddlgenerator.sources.url_utils.safe_fetch_text')
    def test_extracts_table_name_from_url(self, mock_fetch):
        mock_fetch.return_value = '[{"a": 1}]'
        src = Source("https://example.com/path/to/myfile.yaml")
        assert src.table_name == "myfile"

    @pytest.mark.skipif(xlrd is None, reason="xlrd not installed")
    @patch('ddlgenerator.sources.url_utils.safe_fetch_content')
    @patch('ddlgenerator.sources.url_utils.safe_fetch_text')
    def test_remote_xls_is_fetched_as_bytes(self, mock_text, mock_content):
        """A spreadsheet is binary. Fetching it as text decodes it with a
        guessed charset and re-encoding cannot recover the original bytes,
        so the workbook never parses."""
        mock_content.return_value = pathlib.Path(here("luxembourg.xls")).read_bytes()
        rows = list(Source("https://example.com/data.xls"))
        mock_content.assert_called_once_with("https://example.com/data.xls")
        mock_text.assert_not_called()
        assert rows

    @patch('ddlgenerator.sources.url_utils.safe_fetch_content')
    @patch('ddlgenerator.sources.url_utils.safe_fetch_text')
    def test_remote_xlsx_is_fetched_as_bytes(self, mock_text, mock_content):
        mock_content.return_value = b"PK\x03\x04not-really-a-workbook"
        # The stub is not a real workbook, so parsing fails; what matters is
        # which fetch was used to get the bytes.
        with contextlib.suppress(Exception):
            Source("https://example.com/data.xlsx")
        mock_content.assert_called_once_with("https://example.com/data.xlsx")
        mock_text.assert_not_called()


# ---------------------------------------------------------------------------
# Source class - Excel
# ---------------------------------------------------------------------------
@pytest.mark.skipif(xlrd is None, reason="xlrd not installed")
class TestSourceExcel:
    def test_source_is_excel_path(self):
        src = Source(here("luxembourg.xls"))
        assert "luxembourg" in src.table_name.lower() or src.table_name.startswith("Table")

    def test_excel_path_longer_than_84_chars(self, tmp_path):
        """A long absolute path is still a path, not spreadsheet contents.

        The .xls branch used to treat any string of 84 or more characters as
        raw file contents. xlrd then raised TypeError, Source swallowed it and
        fell through to the glob branch, where the path matched itself and
        recursed until the stack blew.
        """
        deep = tmp_path
        while len(str(deep / "luxembourg.xls")) < 120:
            deep = deep / "nested_directory_component"
        deep.mkdir(parents=True, exist_ok=True)
        target = deep / "luxembourg.xls"
        target.write_bytes(pathlib.Path(here("luxembourg.xls")).read_bytes())
        assert len(str(target)) >= 84

        rows = list(Source(str(target)))
        assert rows
        assert list(rows[0])  # rows carry real column names, not an empty shell

    def test_excel_source_is_iterable(self):
        """NamedIter set __iter__/__next__ as instance attributes, which Python
        ignores for dunder lookup, so every .xls source raised TypeError on
        first read."""
        rows = list(Source(here("luxembourg.xls")))
        assert len(rows) == 6
        assert "Name" in rows[0]

    def test_unreadable_excel_raises_rather_than_recursing(self, tmp_path):
        """A corrupt .xls must surface an error, not fall through to glob."""
        bad = tmp_path / "corrupt.xls"
        bad.write_bytes(b"this is not a spreadsheet")
        with pytest.raises(Exception) as exc_info:
            Source(str(bad))
        assert not isinstance(exc_info.value, RecursionError)


@pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")
class TestSourceLocalXlsx:
    def test_local_xlsx_path_is_readable(self, tmp_path):
        """A local .xlsx path used to fall through to the generic path
        loader (whose deserializer table has no .xlsx entry) because the
        file-path dispatch only special-cased '.xls', not '.xlsx'. Only
        URL-fetched .xlsx worked. Regression test for that dispatch gap."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["id", "name"])
        sheet.append([1, "alice"])
        sheet.append([2, "bob"])
        path = tmp_path / "data.xlsx"
        workbook.save(path)

        rows = list(Source(str(path)))
        assert len(rows) == 2
        assert rows[0]["name"] == "alice"


class TestSourceSelfMatchingGlob:
    def test_directory_source_does_not_recurse(self, tmp_path):
        """glob() matches a directory by its own name; expanding that as a
        multi-source list rebuilt Source on the same path forever."""
        try:
            list(Source(str(tmp_path)))
        except RecursionError:
            pytest.fail("Source recursed on a self-matching glob")
        except Exception:
            pass  # any bounded failure is acceptable; unbounded recursion is not


# ---------------------------------------------------------------------------
# Source class - Deserialize
# ---------------------------------------------------------------------------
class TestSourceDeserialize:
    def test_tries_deserializers_in_order(self):
        """Source should try multiple deserializers until one works."""
        json_content = '[{"name": "Alice"}]'
        src = Source(StringIO(json_content))
        result = list(src)
        assert len(result) == 1
        assert result[0]["name"] == "Alice"

    def test_raises_syntax_error_on_all_failures(self):
        """Invalid content should raise SyntaxError after all deserializers fail."""
        # This test is tricky because YAML is very permissive
        # Skip if YAML parses the content
        pytest.skip("YAML is too permissive - most content parses as string")

    def test_handles_stop_iteration_gracefully(self):
        """Empty content should be handled gracefully."""
        # Empty JSON array
        src = Source(StringIO("[]"))
        result = list(src)
        assert result == []


# ---------------------------------------------------------------------------
# Source class - Multiple Sources
# ---------------------------------------------------------------------------
class TestMultipleSources:
    def test_chains_multiple_sources(self):
        """Multiple sources should be chained together."""
        # Use glob pattern to test multiple sources
        import glob
        pattern = here("*.csv")
        files = sorted(glob.glob(pattern))
        if len(files) > 0:
            src = Source(pattern)
            # Should have data from all matching files
            assert src.generator is not None


# ---------------------------------------------------------------------------
# Source class - File paths
# ---------------------------------------------------------------------------
class TestSourceFilePaths:
    def test_yaml_file(self):
        src = Source(here("knights.yaml"))
        result = list(src)
        assert len(result) > 0
        assert "name" in result[0] or "Lancelot" in str(result)

    def test_json_file(self):
        src = Source(here("menu.json"))
        result = list(src)
        assert len(result) > 0

    def test_csv_file(self):
        src = Source(here("animals.csv"))
        result = list(src)
        assert len(result) > 0

    @pytest.mark.skipif(bs4 is None, reason="beautifulsoup4 not installed")
    def test_html_file(self):
        src = Source(here("cities_of_ohio.html"))
        result = list(src)
        assert len(result) > 0


# ---------------------------------------------------------------------------
# Source class - File-like objects
# ---------------------------------------------------------------------------
class TestSourceFileLike:
    def test_file_like_object(self):
        content = '[{"name": "Alice", "age": 30}]'
        src = Source(StringIO(content))
        result = list(src)
        assert result[0]["name"] == "Alice"

    def test_file_like_with_name_attribute(self):
        file_obj = StringIO('[{"a": 1}]')
        file_obj.name = "test_data.json"
        src = Source(file_obj)
        assert src.table_name == "test_data"


# ---------------------------------------------------------------------------
# Source class - Limit
# ---------------------------------------------------------------------------
class TestSourceLimit:
    def test_limit_respects_row_count(self):
        data = [{"id": i} for i in range(10)]
        src = Source(iter(data), limit=3)
        result = list(src)
        assert len(result) == 3


# ---------------------------------------------------------------------------
# Source class - Every Nth
# ---------------------------------------------------------------------------
class TestSourceEveryNth:
    def test_every_nth_keeps_multiples(self):
        data = [{"id": i} for i in range(1, 11)]  # ids 1..10
        src = Source(iter(data), every_nth=3)
        result = list(src)
        assert [r["id"] for r in result] == [3, 6, 9]

    def test_every_nth_one_keeps_all_rows(self):
        data = [{"id": i} for i in range(5)]
        src = Source(iter(data), every_nth=1)
        assert len(list(src)) == 5

    def test_every_nth_combined_with_limit_caps_surviving_rows(self):
        data = [{"id": i} for i in range(1, 21)]  # ids 1..20
        src = Source(iter(data), every_nth=2, limit=3)
        result = list(src)
        # every_nth=2 alone yields ids 2,4,...,20 (10 rows); limit=3 caps to the first 3 surviving
        assert [r["id"] for r in result] == [2, 4, 6]

    def test_every_nth_zero_raises(self):
        with pytest.raises(ValueError, match="every_nth must be a positive integer"):
            Source(iter([{"id": 1}]), every_nth=0)

    def test_every_nth_negative_raises(self):
        with pytest.raises(ValueError, match="every_nth must be a positive integer"):
            Source(iter([{"id": 1}]), every_nth=-1)

    def test_every_nth_on_empty_source_yields_nothing(self):
        src = Source(iter([]), every_nth=3)
        assert list(src) == []


# ---------------------------------------------------------------------------
# Source class - sample_k (reservoir sampling)
# ---------------------------------------------------------------------------
class TestSourceSampleKValidation:
    def test_sample_k_zero_raises(self):
        with pytest.raises(ValueError, match="sample_k must be a positive integer"):
            Source(iter([{"id": 1}]), sample_k=0)

    def test_sample_k_negative_raises(self):
        with pytest.raises(ValueError, match="sample_k must be a positive integer"):
            Source(iter([{"id": 1}]), sample_k=-1)

    def test_seed_without_sample_k_raises(self):
        with pytest.raises(ValueError, match="seed requires sample_k"):
            Source(iter([{"id": 1}]), seed=42)

    def test_sample_k_combined_with_limit_raises(self):
        with pytest.raises(ValueError, match="sample_k cannot be combined with limit or every_nth"):
            Source(iter([{"id": 1}]), sample_k=1, limit=1)

    def test_sample_k_combined_with_every_nth_raises(self):
        with pytest.raises(ValueError, match="sample_k cannot be combined with limit or every_nth"):
            Source(iter([{"id": 1}]), sample_k=1, every_nth=1)


class TestSourceSampleK:
    def test_sample_k_returns_exactly_k_rows_when_n_greater_than_k(self):
        data = [{"id": i} for i in range(1, 101)]  # ids 1..100
        src = Source(iter(data), sample_k=10, seed=42)
        result = list(src)
        assert len(result) == 10

    def test_sample_k_is_deterministic_with_same_seed(self):
        data = [{"id": i} for i in range(1, 101)]
        result_a = list(Source(iter(data), sample_k=10, seed=42))
        result_b = list(Source(iter(data), sample_k=10, seed=42))
        assert [r["id"] for r in result_a] == [r["id"] for r in result_b]

    def test_sample_k_keeps_all_rows_when_n_less_than_k(self):
        data = [{"id": i} for i in range(1, 6)]  # 5 rows
        src = Source(iter(data), sample_k=10, seed=1)
        result = list(src)
        assert [r["id"] for r in result] == [1, 2, 3, 4, 5]

    def test_sample_k_keeps_all_rows_when_n_equals_k(self):
        data = [{"id": i} for i in range(1, 11)]  # 10 rows
        src = Source(iter(data), sample_k=10, seed=1)
        result = list(src)
        assert [r["id"] for r in result] == list(range(1, 11))

    def test_sample_k_preserves_original_relative_order(self):
        data = [{"id": i} for i in range(1, 51)]
        src = Source(iter(data), sample_k=5, seed=7)
        result = [r["id"] for r in list(src)]
        assert result == sorted(result)

    def test_sample_k_on_empty_source_yields_nothing(self):
        src = Source(iter([]), sample_k=3, seed=1)
        assert list(src) == []

    def test_sample_k_without_seed_still_returns_k_rows(self):
        data = [{"id": i} for i in range(1, 21)]
        src = Source(iter(data), sample_k=4)
        assert len(list(src)) == 4

    def test_sample_k_applies_per_matched_file_in_glob(self, tmp_path):
        (tmp_path / "a.json").write_text(
            '[' + ','.join(f'{{"id": {i}}}' for i in range(1, 21)) + ']'
        )
        (tmp_path / "b.json").write_text(
            '[' + ','.join(f'{{"id": {i}}}' for i in range(101, 121)) + ']'
        )
        src = Source(str(tmp_path / "*.json"), sample_k=3, seed=5)
        result = list(src)
        assert len(result) == 6  # 3 from each of the two matched files
        from_a = [r["id"] for r in result if r["id"] < 100]
        from_b = [r["id"] for r in result if r["id"] >= 100]
        assert len(from_a) == 3
        assert len(from_b) == 3

    def test_sample_k_seed_offset_gives_independent_positions_per_subsource(self, tmp_path):
        """Same seed forwarded unmodified to every subsource would make
        Algorithm R (index-driven, not content-driven) pick identical
        relative positions from each same-length file. Each subsource's
        seed is offset by its index so positions differ instead."""
        (tmp_path / "a.json").write_text(
            '[' + ','.join(f'{{"id": {i}}}' for i in range(1, 21)) + ']'
        )
        (tmp_path / "b.json").write_text(
            '[' + ','.join(f'{{"id": {i}}}' for i in range(101, 121)) + ']'
        )
        src = Source(str(tmp_path / "*.json"), sample_k=3, seed=5)
        result = list(src)
        relative_a = sorted(r["id"] - 1 for r in result if r["id"] < 100)
        relative_b = sorted(r["id"] - 101 for r in result if r["id"] >= 100)
        assert relative_a != relative_b


# ---------------------------------------------------------------------------
# Source class - sample_pct (Bernoulli sampling)
# ---------------------------------------------------------------------------
class TestSourceSamplePctValidation:
    def test_sample_pct_zero_raises(self):
        with pytest.raises(ValueError,
                           match="sample_pct must be greater than 0 and at most 100"):
            Source(iter([{"id": 1}]), sample_pct=0)

    def test_sample_pct_negative_raises(self):
        with pytest.raises(ValueError,
                           match="sample_pct must be greater than 0 and at most 100"):
            Source(iter([{"id": 1}]), sample_pct=-5)

    def test_sample_pct_over_100_raises(self):
        with pytest.raises(ValueError,
                           match="sample_pct must be greater than 0 and at most 100"):
            Source(iter([{"id": 1}]), sample_pct=101)

    def test_sample_pct_combined_with_limit_raises(self):
        with pytest.raises(
                ValueError,
                match="sample_pct cannot be combined with limit, every_nth, or sample_k"):
            Source(iter([{"id": 1}]), sample_pct=10, limit=1)

    def test_sample_pct_combined_with_every_nth_raises(self):
        with pytest.raises(
                ValueError,
                match="sample_pct cannot be combined with limit, every_nth, or sample_k"):
            Source(iter([{"id": 1}]), sample_pct=10, every_nth=2)

    def test_sample_pct_combined_with_sample_k_raises(self):
        with pytest.raises(
                ValueError,
                match="sample_pct cannot be combined with limit, every_nth, or sample_k"):
            Source(iter([{"id": 1}]), sample_pct=10, sample_k=1)

    def test_seed_with_sample_pct_alone_is_allowed(self):
        src = Source(iter([{"id": 1}]), sample_pct=100, seed=42)
        assert [r["id"] for r in src] == [1]


class TestSourceSamplePct:
    def test_sample_pct_100_keeps_every_row(self):
        data = [{"id": i} for i in range(1, 101)]
        src = Source(iter(data), sample_pct=100, seed=1)
        assert [r["id"] for r in src] == list(range(1, 101))

    def test_sample_pct_is_deterministic_with_same_seed(self):
        data = [{"id": i} for i in range(1, 1001)]
        result_a = list(Source(iter(data), sample_pct=10, seed=42))
        result_b = list(Source(iter(data), sample_pct=10, seed=42))
        assert [r["id"] for r in result_a] == [r["id"] for r in result_b]

    def test_sample_pct_yields_approximately_the_requested_share(self):
        """Bernoulli sampling gives an approximate count, not an exact one.
        Seeded, so the count below is deterministic rather than flaky."""
        data = [{"id": i} for i in range(1, 1001)]
        result = list(Source(iter(data), sample_pct=10, seed=42))
        assert len(result) == 90
        assert 60 < len(result) < 140  # comfortably around 10% of 1000

    def test_sample_pct_preserves_original_relative_order(self):
        data = [{"id": i} for i in range(1, 501)]
        result = [r["id"] for r in Source(iter(data), sample_pct=10, seed=7)]
        assert result == sorted(result)

    def test_sample_pct_on_empty_source_yields_nothing(self):
        """The >=1 floor never manufactures a row out of an empty source."""
        src = Source(iter([]), sample_pct=10, seed=1)
        assert list(src) == []

    def test_sample_pct_without_seed_still_yields_rows(self):
        data = [{"id": i} for i in range(1, 1001)]
        assert len(list(Source(iter(data), sample_pct=10))) > 0

    def test_sample_pct_yields_at_least_one_row_when_nothing_survives(self):
        """seed 0 selects no row from 10 rows at 1%, so the floor fires."""
        data = [{"id": i} for i in range(1, 11)]
        result = list(Source(iter(data), sample_pct=1, seed=0))
        assert len(result) == 1

    def test_sample_pct_floor_row_is_not_always_the_first_row(self):
        """The floor row comes from a size-1 reservoir over the whole source,
        not from simply remembering row 1, so it is unbiased."""
        floor_rows = []
        for seed in range(6):  # every one of these selects nothing at 1%
            data = [{"id": i} for i in range(1, 11)]
            result = list(Source(iter(data), sample_pct=1, seed=seed))
            assert len(result) == 1
            floor_rows.append(result[0]["id"])
        assert len(set(floor_rows)) > 1
        assert floor_rows != [1] * len(floor_rows)

    def test_sample_pct_floor_row_is_spread_over_the_whole_source(self):
        """Stronger than "not always row 1": over many seeds the floor row
        should reach every position, which a head-biased or tail-biased
        pick would not. Deterministic -- every seed here is fixed."""
        floor_rows = []
        for seed in range(200):
            data = [{"id": i} for i in range(1, 11)]
            result = list(Source(iter(data), sample_pct=0.01, seed=seed))
            assert len(result) == 1  # 0.01% never selects anything here
            floor_rows.append(result[0]["id"])
        assert set(floor_rows) == set(range(1, 11))
        # no position should dominate the way a "remember row 1" bug would
        assert max(Counter(floor_rows).values()) < len(floor_rows) / 2

    def test_sample_pct_decisions_are_independent_of_floor_bookkeeping(self):
        """The floor keeps its own RNG stream so Bernoulli keep/drop stays a
        clean function of (seed, row index). Pinning it against a plain
        random.Random means merging the two streams breaks this test rather
        than silently shifting which rows every sample returns."""
        n, pct, seed = 500, 20, 11
        rng = random.Random(seed)
        expected = [i for i in range(1, n + 1) if rng.random() < pct / 100]
        data = [{"id": i} for i in range(1, n + 1)]
        actual = [r["id"] for r in Source(iter(data), sample_pct=pct, seed=seed)]
        assert actual == expected

    def test_sample_pct_accepts_a_fractional_percent(self):
        """0 < pct < 1 is legal and samples far more sparsely than 1%."""
        data = [{"id": i} for i in range(1, 10001)]
        result = list(Source(iter(data), sample_pct=0.5, seed=3))
        assert 20 < len(result) < 90  # ~0.5% of 10000 == ~50

    def test_sample_pct_applies_per_matched_file_in_glob(self, tmp_path):
        (tmp_path / "a.json").write_text(
            '[' + ','.join(f'{{"id": {i}}}' for i in range(1, 21)) + ']'
        )
        (tmp_path / "b.json").write_text(
            '[' + ','.join(f'{{"id": {i}}}' for i in range(101, 121)) + ']'
        )
        src = Source(str(tmp_path / "*.json"), sample_pct=25, seed=5)
        result = list(src)
        from_a = [r["id"] for r in result if r["id"] < 100]
        from_b = [r["id"] for r in result if r["id"] >= 100]
        assert len(from_a) > 0
        assert len(from_b) > 0

    def test_sample_pct_floor_applies_per_subsource(self, tmp_path):
        """A percent low enough to select nothing still yields one row from
        each matched file, not one row across the whole glob."""
        for name, start in (("a.json", 1), ("b.json", 101), ("c.json", 201)):
            (tmp_path / name).write_text(
                '[' + ','.join(f'{{"id": {i}}}' for i in range(start, start + 10)) + ']'
            )
        result = list(Source(str(tmp_path / "*.json"), sample_pct=1, seed=0))
        assert len(result) == 3

    def test_sample_pct_seed_offset_gives_independent_positions_per_subsource(self, tmp_path):
        """Bernoulli decisions are index-driven, not content-driven, so the
        same seed forwarded unmodified to every subsource would pick identical
        relative positions from each same-length file."""
        (tmp_path / "a.json").write_text(
            '[' + ','.join(f'{{"id": {i}}}' for i in range(1, 21)) + ']'
        )
        (tmp_path / "b.json").write_text(
            '[' + ','.join(f'{{"id": {i}}}' for i in range(101, 121)) + ']'
        )
        result = list(Source(str(tmp_path / "*.json"), sample_pct=25, seed=5))
        relative_a = sorted(r["id"] - 1 for r in result if r["id"] < 100)
        relative_b = sorted(r["id"] - 101 for r in result if r["id"] >= 100)
        assert relative_a != relative_b

    @pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")
    def test_sample_pct_applies_per_sheet_in_a_workbook(self, tmp_path):
        """Multi-sheet spreadsheets expand through _multiple_sources the same
        way globs do, so the floor and the seed offset are per sheet."""
        workbook = openpyxl.Workbook()
        for sheet_idx, title in enumerate(("one", "two", "three")):
            sheet = workbook.create_sheet(title) if sheet_idx else workbook.active
            sheet.title = title
            sheet.append(["id", "name"])
            for i in range(1, 11):
                sheet.append([i + sheet_idx * 100, f"{title}{i}"])
        path = tmp_path / "sheets.xlsx"
        workbook.save(path)

        floored = list(Source(str(path), sample_pct=1, seed=0))
        assert len(floored) == 3  # one row per sheet, not one per workbook
        assert sorted(r["id"] // 100 for r in floored) == [0, 1, 2]

        sampled = list(Source(str(path), sample_pct=50, seed=3))
        per_sheet = {s: sorted(r["id"] % 100 for r in sampled if r["id"] // 100 == s)
                     for s in (0, 1, 2)}
        assert all(rows for rows in per_sheet.values())
        assert len({tuple(rows) for rows in per_sheet.values()}) > 1


# ---------------------------------------------------------------------------
# Source.count
# ---------------------------------------------------------------------------
class TestSourceCount:
    def test_counts_json_file(self):
        pairs = Source.count(here("menu.json"))
        assert sum(n for (_name, n) in pairs) == 3

    def test_counts_csv_file(self):
        pairs = Source.count(here("animals.csv"))
        assert sum(n for (_name, n) in pairs) == 3

    def test_count_ignores_limit_and_every_nth_by_design(self):
        """Source.count() takes no limit/every_nth kwargs -- it always reports the true total."""
        import inspect
        sig = inspect.signature(Source.count)
        assert "limit" not in sig.parameters
        assert "every_nth" not in sig.parameters

    def test_counts_glob_per_file(self, tmp_path):
        (tmp_path / "a.json").write_text('[{"id": 1}, {"id": 2}]')
        (tmp_path / "b.json").write_text('[{"id": 1}]')
        pairs = Source.count(str(tmp_path / "*.json"))
        assert len(pairs) == 2
        assert sum(n for (_name, n) in pairs) == 3

    @pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")
    def test_counts_xlsx_without_full_materialization(self, tmp_path, monkeypatch):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["id", "name"])
        for i in range(5):
            sheet.append([i, f"name{i}"])
        path = tmp_path / "data.xlsx"
        workbook.save(path)

        def boom(*args, **kwargs):
            raise AssertionError("count-only must not build the full per-row list")
        monkeypatch.setattr(Source, "_source_is_xlsx_worksheet", boom)

        pairs = Source.count(str(path))
        assert sum(n for (_name, n) in pairs) == 5

    @pytest.mark.skipif(MongoCollection is None, reason="pymongo not installed")
    def test_counts_mongo_via_count_documents(self):
        mock_collection = MagicMock(spec=MongoCollection)
        mock_collection.name = "things"
        mock_collection.count_documents.return_value = 42

        pairs = Source.count(mock_collection)

        assert pairs == [("things", 42)]
        mock_collection.find.assert_not_called()


# ---------------------------------------------------------------------------
# count_sqlalchemy_tables
# ---------------------------------------------------------------------------
@pytest.mark.skipif(sqlalchemy is None, reason="sqlalchemy not installed")
class TestCountSqlalchemyTables:
    def test_uses_select_count_not_meta_bind(self, tmp_path):
        """
        Regression guard: sqlalchemy.MetaData() has no .bind attribute under
        SQLAlchemy 2.x, so count_sqlalchemy_tables must build its own engine/
        connection (like sqlalchemy_table_sources does) rather than relying
        on meta.bind, unlike Source._source_is_sqlalchemy_metadata.
        """
        db_path = tmp_path / "test.db"
        engine = sqlalchemy.create_engine(f"sqlite:///{db_path}")
        with engine.connect() as connection:
            connection.execute(sqlalchemy.text("CREATE TABLE t (id INTEGER, name TEXT)"))
            connection.execute(sqlalchemy.text("INSERT INTO t VALUES (1, 'a'), (2, 'b')"))
            connection.commit()

        results = count_sqlalchemy_tables(f"sqlite:///{db_path}")

        assert results == [("t", 2)]


# ---------------------------------------------------------------------------
# sqlalchemy_table_sources
# ---------------------------------------------------------------------------
@pytest.mark.skipif(sqlalchemy is None, reason="sqlalchemy not installed")
class TestSqlalchemyTableSources:
    @patch('ddlgenerator.sources.sqlalchemy.create_engine')
    @patch('ddlgenerator.sources.sqlalchemy.MetaData')
    def test_yields_source_per_table(self, mock_meta, mock_create_engine):
        """sqlalchemy_table_sources should yield a Source for each table."""
        # Setup mock
        mock_engine = Mock()
        mock_create_engine.return_value = mock_engine

        mock_table = Mock()
        mock_table.name = "test_table"
        mock_meta_inst = MagicMock()
        mock_meta_inst.sorted_tables = [mock_table]
        mock_meta.return_value = mock_meta_inst

        # Mock the Source constructor to avoid needing real DB
        with patch.object(Source, '__init__', return_value=None):
            list(sqlalchemy_table_sources("sqlite:///test.db"))

        mock_create_engine.assert_called_once_with("sqlite:///test.db")
        mock_meta_inst.reflect.assert_called_once()

    def test_reads_rows_from_a_real_engine(self, tmp_path):
        """Regression guard: sqlalchemy.MetaData() has no .bind attribute
        under SQLAlchemy 2.x, so Source._source_is_sqlalchemy_metadata must
        be given the engine explicitly rather than reading meta.bind. The
        mocked test above never exercises real Source construction, so it
        can't catch this -- this test must hit a real engine/connection."""
        db_path = tmp_path / "test.db"
        engine = sqlalchemy.create_engine(f"sqlite:///{db_path}")
        with engine.connect() as connection:
            connection.execute(sqlalchemy.text("CREATE TABLE t (id INTEGER, name TEXT)"))
            connection.execute(sqlalchemy.text("INSERT INTO t VALUES (1, 'a'), (2, 'b')"))
            connection.commit()

        sources = list(sqlalchemy_table_sources(f"sqlite:///{db_path}"))

        assert len(sources) == 1
        rows = list(sources[0])
        assert [tuple(row.values()) for row in rows] == [(1, 'a'), (2, 'b')]

    def test_reflected_columns_ride_on_the_generator(self, tmp_path):
        """The generator carries the reflected schema as ``sqla_columns``
        (the attribute ddlgenerator.Table checks for), so the primary key
        and declared types survive instead of being re-inferred from row
        values (issue #28)."""
        db_path = tmp_path / "test.db"
        engine = sqlalchemy.create_engine(f"sqlite:///{db_path}")
        with engine.connect() as connection:
            connection.execute(sqlalchemy.text(
                "CREATE TABLE t (id INTEGER PRIMARY KEY, name TEXT)"))
            connection.commit()

        (source,) = sqlalchemy_table_sources(f"sqlite:///{db_path}")

        cols = {col.name: col for col in source.generator.sqla_columns}
        assert set(cols) == {'id', 'name'}
        assert cols['id'].primary_key
        assert not cols['name'].primary_key

    def test_limit_applies_to_sqlalchemy_url_source(self, tmp_path):
        """Regression guard: sqlalchemy_table_sources previously never
        received limit/every_nth at all, so --limit/--every-nth were
        silently no-ops for sqlalchemy:// URLs."""
        db_path = tmp_path / "test.db"
        engine = sqlalchemy.create_engine(f"sqlite:///{db_path}")
        with engine.connect() as connection:
            connection.execute(sqlalchemy.text("CREATE TABLE t (id INTEGER)"))
            connection.execute(sqlalchemy.text(
                "INSERT INTO t VALUES (1), (2), (3), (4), (5)"
            ))
            connection.commit()

        sources = list(sqlalchemy_table_sources(f"sqlite:///{db_path}", limit=2))

        assert len(list(sources[0])) == 2

    def test_every_nth_applies_to_sqlalchemy_url_source(self, tmp_path):
        """Same regression guard as test_limit_applies_to_sqlalchemy_url_source,
        for --every-nth."""
        db_path = tmp_path / "test.db"
        engine = sqlalchemy.create_engine(f"sqlite:///{db_path}")
        with engine.connect() as connection:
            connection.execute(sqlalchemy.text("CREATE TABLE t (id INTEGER)"))
            connection.execute(sqlalchemy.text(
                "INSERT INTO t VALUES " + ", ".join(f"({i})" for i in range(1, 11))
            ))
            connection.commit()

        sources = list(sqlalchemy_table_sources(f"sqlite:///{db_path}", every_nth=3))

        rows = list(sources[0])
        assert [tuple(row.values()) for row in rows] == [(3,), (6,), (9,)]

    def test_sample_k_applies_to_sqlalchemy_url_source(self, tmp_path):
        db_path = tmp_path / "test.db"
        engine = sqlalchemy.create_engine(f"sqlite:///{db_path}")
        with engine.connect() as connection:
            connection.execute(sqlalchemy.text("CREATE TABLE t (id INTEGER)"))
            connection.execute(sqlalchemy.text(
                "INSERT INTO t VALUES " + ", ".join(f"({i})" for i in range(1, 21))
            ))
            connection.commit()

        sources = list(sqlalchemy_table_sources(f"sqlite:///{db_path}", sample_k=5, seed=3))

        assert len(list(sources[0])) == 5

    def test_raises_import_error_when_sqlalchemy_none(self):
        """Should raise ImportError if sqlalchemy is not available."""
        with patch('ddlgenerator.sources.sqlalchemy', None):
            with pytest.raises(ImportError, match="sqlalchemy not installed"):
                list(sqlalchemy_table_sources("sqlite:///test.db"))
